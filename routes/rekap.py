from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from models import Santri, Acara, Absensi
from extensions import db
import io, openpyxl, logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

logger = logging.getLogger(__name__)
rekap_bp = Blueprint('rekap', __name__)

PER_PAGE = 100   # paginasi rekap


def get_filtered(acara_id, kelas, tanggal):
    q = db.session.query(Absensi, Santri, Acara)\
          .join(Santri, Absensi.santri_id == Santri.id)\
          .join(Acara,  Absensi.acara_id  == Acara.id)
    if acara_id:
        q = q.filter(Absensi.acara_id == acara_id)
    if kelas:
        q = q.filter(Santri.kelas == kelas)
    if tanggal:
        from datetime import date
        q = q.filter(Acara.tanggal == date.fromisoformat(tanggal))
    return q.order_by(Absensi.waktu.desc())


@rekap_bp.route('/rekap')
@login_required
def index():
    acara_id  = request.args.get('acara_id', type=int)
    kelas     = request.args.get('kelas', '')
    tanggal   = request.args.get('tanggal', '')
    page      = request.args.get('page', 1, type=int)

    q         = get_filtered(acara_id, kelas, tanggal)
    total     = q.count()
    rows      = q.offset((page - 1) * PER_PAGE).limit(PER_PAGE).all()

    tepat     = db.session.query(Absensi).join(Santri, Absensi.santri_id==Santri.id)\
                    .join(Acara, Absensi.acara_id==Acara.id)
    if acara_id: tepat = tepat.filter(Absensi.acara_id == acara_id)
    if kelas:    tepat = tepat.filter(Santri.kelas == kelas)
    tepat_n   = tepat.filter(Absensi.status=='tepat_waktu').count()
    terlambat_n = total - tepat_n

    acara_list  = Acara.query.order_by(Acara.dibuat.desc()).all()
    kelas_list  = sorted([k[0] for k in db.session.query(Santri.kelas).distinct().all() if k[0]])
    pages       = (total + PER_PAGE - 1) // PER_PAGE

    return render_template('rekap.html',
        rows=rows, tepat=tepat_n, terlambat=terlambat_n,
        total=total, page=page, pages=pages,
        acara_id=acara_id, kelas=kelas, tanggal=tanggal,
        acara_list=acara_list, kelas_list=kelas_list
    )


@rekap_bp.route('/rekap/excel')
@login_required
def excel():
    acara_id = request.args.get('acara_id', type=int)
    kelas    = request.args.get('kelas', '')
    tanggal  = request.args.get('tanggal', '')
    rows     = get_filtered(acara_id, kelas, tanggal).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rekap Absensi'
    hfill = PatternFill('solid', fgColor='2563EB')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'),  bottom=Side(style='thin'))
    ws.merge_cells('A1:H1')
    ws['A1'] = 'REKAP ABSENSI SANTRI'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Dicetak: {datetime.now().strftime("%d %B %Y, %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    headers = ['No','Nama Santri','NIS','Kelas','Orang Tua','Nama Acara','Waktu Absensi','Status']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = thin
        c.alignment = Alignment(horizontal='center')
    for i, (ab, s, a) in enumerate(rows, 1):
        row_data = [i, s.nama, s.nis, s.kelas or '-', s.orang_tua or '-',
                    a.nama, ab.waktu.strftime('%d/%m/%Y %H:%M'),
                    'Tepat Waktu' if ab.status == 'tepat_waktu' else 'Terlambat']
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i+4, column=col, value=val)
            c.border = thin
            if col == 8:
                c.font = Font(color='16A34A' if ab.status == 'tepat_waktu' else 'DC2626', bold=True)
    for col, w in enumerate([5,25,15,12,20,25,20,14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'rekap_absensi_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rekap_bp.route('/rekap/pdf')
@login_required
def pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    acara_id = request.args.get('acara_id', type=int)
    kelas    = request.args.get('kelas', '')
    tanggal  = request.args.get('tanggal', '')
    rows     = get_filtered(acara_id, kelas, tanggal).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles   = getSampleStyleSheet()
    elements = [
        Paragraph('Rekap Absensi Santri',
                  ParagraphStyle('T', fontSize=16, fontName='Helvetica-Bold', spaceAfter=4)),
        Paragraph(f'Dicetak: {datetime.now().strftime("%d %B %Y, %H:%M")}',
                  ParagraphStyle('S', fontSize=9, textColor=colors.grey, spaceAfter=12)),
    ]
    tdata = [['No','Nama Santri','NIS','Kelas','Nama Acara','Tanggal','Waktu','Status']]
    for i, (ab, s, a) in enumerate(rows, 1):
        tdata.append([str(i), s.nama, s.nis, s.kelas or '-',
                      a.nama, a.tanggal.strftime('%d/%m/%Y'),
                      ab.waktu.strftime('%H:%M'),
                      'Tepat' if ab.status == 'tepat_waktu' else 'Terlambat'])
    t = Table(tdata, colWidths=[1.2*cm,5.5*cm,3.5*cm,2.5*cm,5.5*cm,3*cm,2.5*cm,3*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#2563EB')),
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#F1F5F9')]),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'rekap_absensi_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')
