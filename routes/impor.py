from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from extensions import db
from models import Santri
import io, openpyxl, logging

logger = logging.getLogger(__name__)

impor_bp = Blueprint('impor', __name__)

# Magic bytes untuk validasi konten file (bukan hanya ekstensi)
_XLSX_MAGIC = b'PK\x03\x04'   # ZIP (xlsx/xls baru)
_XLS_MAGIC  = b'\xd0\xcf\x11\xe0'  # Compound Document (xls lama)
_CSV_UTF8   = None  # CSV tidak punya magic bytes — cukup cek ekstensi + decode

def _is_valid_spreadsheet(content: bytes, ext: str) -> bool:
    """Pastikan isi file sesuai dengan ekstensi yang diklaim."""
    if ext == 'csv':
        # CSV: tidak ada magic bytes, tapi coba decode sebagai teks
        try:
            content[:1024].decode('utf-8')
            return True
        except UnicodeDecodeError:
            try:
                content[:1024].decode('latin-1')
                return True
            except Exception:
                return False
    elif ext == 'xlsx':
        return content[:4] == _XLSX_MAGIC
    elif ext == 'xls':
        return content[:4] in (_XLS_MAGIC, _XLSX_MAGIC)
    return False


@impor_bp.route('/import')
@login_required
def index():
    return render_template('import.html')


@impor_bp.route('/import/template')
@login_required
def template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template Santri'

    from openpyxl.styles import Font, PatternFill, Alignment
    hfill = PatternFill('solid', fgColor='2563EB')
    hfont = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(['NIS','Nama Santri','Nama Orang Tua','Alamat','Kelas'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center')

    for r, row in enumerate([
        ['PST2024010','Contoh Nama Santri','Bapak Contoh','Jl. Contoh No. 1, Kota','Kelas 1A'],
        ['PST2024011','Santri Kedua',      'Ibu Contoh',  'Jl. Lainnya No. 5, Kota','Kelas 2B'],
    ], 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    for col, w in enumerate([18,28,24,35,12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='template_import_santri.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@impor_bp.route('/import/upload', methods=['POST'])
@login_required
def upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Pilih file terlebih dahulu!', 'danger')
        return redirect(url_for('impor.index'))

    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls', 'csv'):
        flash('Format file tidak didukung. Gunakan .xlsx, .xls, atau .csv', 'danger')
        return redirect(url_for('impor.index'))

    try:
        content = file.read()

        # Validasi magic bytes — pastikan isi file sesuai ekstensi
        if not _is_valid_spreadsheet(content, ext):
            flash('File tidak valid atau tidak sesuai formatnya.', 'danger')
            return redirect(url_for('impor.index'))

        if ext == 'csv':
            import csv, codecs
            reader   = csv.reader(codecs.iterdecode(io.BytesIO(content), 'utf-8'))
            all_rows = list(reader)
        else:
            wb       = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws       = wb.active
            all_rows = [[str(cell.value or '').strip() for cell in row] for row in ws.iter_rows()]

        if len(all_rows) < 2:
            flash('File kosong atau tidak valid!', 'danger')
            return redirect(url_for('impor.index'))

        valid, errors = [], []
        for i, row in enumerate(all_rows[1:]):
            while len(row) < 5:
                row.append('')
            nis, nama, ot, alamat, kelas = [str(v).strip() for v in row[:5]]
            if not nis and not nama:
                continue
            if not nis or not nama:
                errors.append({'baris': i+2, 'nis': nis, 'nama': nama, 'err': 'NIS dan Nama wajib diisi'})
                continue
            if Santri.query.filter_by(nis=nis).first():
                errors.append({'baris': i+2, 'nis': nis, 'nama': nama, 'err': f'NIS {nis} sudah terdaftar'})
                continue
            valid.append({'nis': nis, 'nama': nama, 'orang_tua': ot, 'alamat': alamat, 'kelas': kelas})

        return render_template('import.html', valid=valid, errors=errors, filename=file.filename)

    except Exception as e:
        logger.error('Import file error: %s', e, exc_info=True)
        flash('Gagal membaca file. Pastikan file tidak rusak dan sesuai format yang didukung.', 'danger')
        return redirect(url_for('impor.index'))


@impor_bp.route('/import/simpan', methods=['POST'])
@login_required
def simpan():
    count = int(request.form.get('count', 0))
    added = 0
    for i in range(count):
        nis    = request.form.get(f'nis_{i}', '').strip()
        nama   = request.form.get(f'nama_{i}', '').strip()
        ot     = request.form.get(f'orang_tua_{i}', '').strip()
        alamat = request.form.get(f'alamat_{i}', '').strip()
        kelas  = request.form.get(f'kelas_{i}', '').strip()
        if nis and nama and not Santri.query.filter_by(nis=nis).first():
            db.session.add(Santri(nis=nis, nama=nama, orang_tua=ot, alamat=alamat, kelas=kelas))
            added += 1
    db.session.commit()
    flash(f'{added} data santri berhasil diimport!', 'success')
    return redirect(url_for('santri.index'))
